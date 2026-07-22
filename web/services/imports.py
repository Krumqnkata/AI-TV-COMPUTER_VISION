"""Previewable CSV/XLSX timetable imports with row-level validation."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from sqlalchemy.orm import Session

from engine.admin_models import ImportRowError, ScheduleImportJob, StaffAccount
from engine.db import Person, Timetable, now_bg
from utils.config import Config
from web.services.admin_control import audit_event


HEADER_ALIASES = {
    "full_name": {"full_name", "person", "person_name", "име", "пълно име", "ученик/учител"},
    "date": {"date", "дата"},
    "period": {"period", "час", "номер на час"},
    "start_time": {"start_time", "start", "начало", "начален час"},
    "end_time": {"end_time", "end", "край", "краен час"},
    "subject": {"subject", "предмет"},
    "class_name": {"class_name", "class", "клас"},
    "room": {"room", "кабинет", "стая"},
}
REQUIRED_FIELDS = ("full_name", "date", "period", "start_time", "end_time", "subject", "room")


@dataclass
class ParsedScheduleRow:
    row_number: int
    person_id: int
    full_name: str
    date: date
    period: int
    start_time: time
    end_time: time
    subject: str
    class_name: str | None
    room: str

    def serializable(self) -> dict[str, Any]:
        result = asdict(self)
        result["date"] = self.date.isoformat()
        result["start_time"] = self.start_time.strftime("%H:%M")
        result["end_time"] = self.end_time.strftime("%H:%M")
        return result


@dataclass
class RowIssue:
    row_number: int
    field_name: str | None
    message: str
    row: dict[str, Any]


@dataclass
class ImportPreview:
    rows: list[ParsedScheduleRow]
    issues: list[RowIssue]

    @property
    def total_rows(self) -> int:
        return len(self.rows) + len({issue.row_number for issue in self.issues})


def _canonical_headers(headers: Iterable[Any]) -> dict[int, str]:
    aliases = {alias.casefold(): canonical for canonical, values in HEADER_ALIASES.items() for alias in values}
    result: dict[int, str] = {}
    for index, raw in enumerate(headers):
        normalized = str(raw or "").strip().casefold()
        if normalized in aliases:
            result[index] = aliases[normalized]
    return result


def _csv_rows(content: bytes) -> list[list[Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1251")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def _xlsx_rows(content: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency error is explicit
        raise RuntimeError("Липсва зависимостта openpyxl за XLSX импорт") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _raw_rows(file_name: str, content: bytes) -> list[dict[str, Any]]:
    if len(content) > Config.MAX_IMPORT_BYTES:
        raise ValueError("Файлът надвишава позволения размер")
    extension = Path(file_name).suffix.lower()
    if extension == ".csv":
        rows = _csv_rows(content)
    elif extension in {".xlsx", ".xlsm"}:
        rows = _xlsx_rows(content)
    else:
        raise ValueError("Поддържат се само CSV и XLSX файлове")
    if not rows:
        raise ValueError("Файлът е празен")
    header_map = _canonical_headers(rows[0])
    missing = [field for field in REQUIRED_FIELDS if field not in header_map.values()]
    if missing:
        raise ValueError("Липсват задължителни колони: " + ", ".join(missing))
    result = []
    for values in rows[1:]:
        row = {field: values[index] if index < len(values) else None for index, field in header_map.items()}
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Използвайте дата във формат ГГГГ-ММ-ДД или ДД.ММ.ГГГГ")


def _parse_time(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    text = str(value or "").strip()
    for pattern in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, pattern).time()
        except ValueError:
            continue
    raise ValueError("Използвайте час във формат ЧЧ:ММ")


def preview_schedule_import(db: Session, file_name: str, content: bytes) -> ImportPreview:
    raw_rows = _raw_rows(file_name, content)
    people_by_name: dict[str, list[Person]] = {}
    for person in db.query(Person).filter(Person.active.is_(True)).all():
        people_by_name.setdefault(person.full_name.strip().casefold(), []).append(person)

    parsed: list[ParsedScheduleRow] = []
    issues: list[RowIssue] = []
    seen: set[tuple[int, date, int]] = set()
    for row_number, raw in enumerate(raw_rows, start=2):
        row_issues: list[RowIssue] = []
        name = str(raw.get("full_name") or "").strip()
        candidates = people_by_name.get(name.casefold(), [])
        if not candidates:
            row_issues.append(RowIssue(row_number, "full_name", "Няма активен човек с това име", raw))
        elif len(candidates) > 1:
            row_issues.append(RowIssue(row_number, "full_name", "Името не е еднозначно", raw))
        try:
            parsed_date = _parse_date(raw.get("date"))
        except ValueError as exc:
            parsed_date = None
            row_issues.append(RowIssue(row_number, "date", str(exc), raw))
        try:
            period = int(raw.get("period"))
            if not 1 <= period <= 20:
                raise ValueError
        except (TypeError, ValueError):
            period = 0
            row_issues.append(RowIssue(row_number, "period", "Номерът на часа трябва да е между 1 и 20", raw))
        try:
            start = _parse_time(raw.get("start_time"))
        except ValueError as exc:
            start = None
            row_issues.append(RowIssue(row_number, "start_time", str(exc), raw))
        try:
            end = _parse_time(raw.get("end_time"))
        except ValueError as exc:
            end = None
            row_issues.append(RowIssue(row_number, "end_time", str(exc), raw))
        if start and end and end <= start:
            row_issues.append(RowIssue(row_number, "end_time", "Краят трябва да е след началото", raw))
        subject = str(raw.get("subject") or "").strip()
        room = str(raw.get("room") or "").strip()
        if not subject:
            row_issues.append(RowIssue(row_number, "subject", "Предметът е задължителен", raw))
        if not room:
            row_issues.append(RowIssue(row_number, "room", "Кабинетът е задължителен", raw))
        if row_issues:
            issues.extend(row_issues)
            continue
        person = candidates[0]
        unique_key = (person.id, parsed_date, period)
        if unique_key in seen:
            issues.append(RowIssue(row_number, None, "Дублиран човек, дата и час във файла", raw))
            continue
        seen.add(unique_key)
        parsed.append(ParsedScheduleRow(
            row_number=row_number,
            person_id=person.id,
            full_name=person.full_name,
            date=parsed_date,
            period=period,
            start_time=start,
            end_time=end,
            subject=subject[:100],
            class_name=str(raw.get("class_name") or "").strip()[:10] or None,
            room=room[:50],
        ))
    return ImportPreview(parsed, issues)


def execute_schedule_import(
    db: Session,
    file_name: str,
    content: bytes,
    mode: str,
    actor: StaffAccount,
    *,
    allow_partial: bool = False,
    ip_address: str | None = None,
) -> ScheduleImportJob:
    if mode not in {"upsert", "replace_range"}:
        raise ValueError("Невалиден режим на импорт")
    job = ScheduleImportJob(
        file_name=Path(file_name).name[:255],
        mode=mode,
        status="validating",
        created_by_staff_id=actor.id,
    )
    db.add(job)
    db.flush()
    try:
        preview = preview_schedule_import(db, file_name, content)
        job.total_rows = preview.total_rows
        job.valid_rows = len(preview.rows)
        job.invalid_rows = len({issue.row_number for issue in preview.issues})
        for issue in preview.issues:
            db.add(ImportRowError(
                job_id=job.id,
                row_number=issue.row_number,
                field_name=issue.field_name,
                message=issue.message[:500],
                row_json=json.dumps(issue.row, ensure_ascii=False, default=str),
            ))
        if preview.issues and not allow_partial:
            job.status = "rejected"
            job.completed_at = now_bg()
            job.result_json = json.dumps({"reason": "validation_errors"})
            db.commit()
            return job

        deleted = 0
        if mode == "replace_range" and preview.rows:
            person_ids = {row.person_id for row in preview.rows}
            first_date = min(row.date for row in preview.rows)
            last_date = max(row.date for row in preview.rows)
            deleted = db.query(Timetable).filter(
                Timetable.person_id.in_(person_ids),
                Timetable.date >= first_date,
                Timetable.date <= last_date,
            ).delete(synchronize_session=False)

        created = 0
        updated = 0
        for row in preview.rows:
            item = db.query(Timetable).filter(
                Timetable.person_id == row.person_id,
                Timetable.date == row.date,
                Timetable.period == row.period,
            ).first()
            if item is None:
                item = Timetable(person_id=row.person_id, date=row.date, period=row.period)
                db.add(item)
                created += 1
            else:
                updated += 1
            item.start_time = row.start_time
            item.end_time = row.end_time
            item.subject = row.subject
            item.class_name = row.class_name
            item.room = row.room

        job.status = "completed"
        job.completed_at = now_bg()
        job.result_json = json.dumps({"created": created, "updated": updated, "deleted": deleted}, ensure_ascii=False)
        audit_event(
            db,
            "schedule.imported",
            f"Импортирано разписание: {created} нови, {updated} обновени",
            actor=actor,
            entity_type="ScheduleImportJob",
            entity_id=job.id,
            changes={"mode": mode, "created": created, "updated": updated, "deleted": deleted, "invalid": job.invalid_rows},
            ip_address=ip_address,
        )
        db.commit()
        db.refresh(job)
        return job
    except Exception as exc:
        job.status = "failed"
        job.completed_at = now_bg()
        job.result_json = json.dumps({"error": str(exc)[:500]}, ensure_ascii=False)
        db.commit()
        raise


def timetable_csv_template() -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["full_name", "date", "period", "start_time", "end_time", "subject", "class_name", "room"])
    writer.writerow(["Име от системата", date.today().isoformat(), 1, "08:00", "08:40", "Математика", "8А", "304"])
    return ("\ufeff" + output.getvalue()).encode("utf-8")
